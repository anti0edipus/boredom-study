import uuid
import json
import random
from datetime import datetime, timezone
from functools import wraps

from flask import (
    Flask, request, session, redirect, url_for,
    render_template, g, abort, send_file
)
import openpyxl
from openpyxl.utils import get_column_letter
import io

from config import Config
import database as db_module
from randomizer import assign_condition

app = Flask(__name__)
app.config.from_object(Config)
app.jinja_env.globals['enumerate'] = enumerate

# ── Page order used for progress display ──────────────────────────────────────
PAGE_STEPS = ['consent', 'demographics', 'ladders', 'traits', 'writing',
              'mancheck', 'boring_task', 'outcome', 'debrief']
STEP_LABELS = {
    'consent':      'Consent',
    'demographics': 'Page 1',
    'ladders':      'Page 2',
    'traits':       'Page 3',
    'writing':      'Page 4',
    'mancheck':     'Page 5',
    'boring_task':  'Page 6',
    'outcome':      'Page 7',
    'debrief':      'Page 8',
}
NEXT_PAGE = {
    'consent':      'demographics',
    'demographics': 'ladders',
    'ladders':      'traits',
    'traits':       'writing',
    'writing':      'mancheck',
    'mancheck':     'boring_task',
    'boring_task':  'outcome',
    'outcome':      'debrief',
    'debrief':      'completed',
}


def utcnow():
    return datetime.now(timezone.utc).isoformat()


# ── DB connection per request ─────────────────────────────────────────────────
@app.before_request
def open_db():
    g.db = db_module.get_db()


@app.teardown_appcontext
def close_db(exc):
    db = g.pop('db', None)
    if db is not None:
        db.close()


# ── Progress helper ────────────────────────────────────────────────────────────
def progress(current_step):
    visible = ['consent', 'demographics', 'ladders', 'traits', 'writing', 'mancheck', 'boring_task', 'outcome', 'debrief']
    try:
        idx = visible.index(current_step if current_step in visible else 'demographics')
    except ValueError:
        idx = 0
    return {'step': idx + 1, 'total': len(visible), 'label': STEP_LABELS.get(current_step, '')}


# ── Participant guard ──────────────────────────────────────────────────────────
def get_current_participant():
    pid = session.get('participant_id')
    if not pid:
        return None
    return db_module.get_participant(g.db, pid)


def require_step(step):
    """Decorator: ensure participant exists and is at or past the required step."""
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            p = get_current_participant()
            if p is None:
                return redirect(url_for('entry'))
            if p['completed']:
                return redirect(url_for('already_done'))
            # If participant is ahead of this step, let them continue from where they are
            current = p['current_step']
            if current != step and PAGE_STEPS.index(current) > PAGE_STEPS.index(step):
                return redirect(url_for(current))
            return f(*args, **kwargs)
        return wrapped
    return decorator


# ── 0. Entry ──────────────────────────────────────────────────────────────────
@app.route('/')
def entry():
    prolific_pid = request.args.get('PROLIFIC_PID', '')
    study_id = request.args.get('STUDY_ID', '')
    session_id_param = request.args.get('SESSION_ID', '')
    test_mode = 1 if not prolific_pid else 0

    # Resume existing session
    if 'participant_id' in session:
        p = get_current_participant()
        if p and not p['completed']:
            return redirect(url_for(p['current_step']))
        elif p and p['completed']:
            return redirect(url_for('already_done'))

    participant_id = str(uuid.uuid4())
    session['participant_id'] = participant_id
    session.permanent = True

    db_module.create_participant(g.db, {
        'participant_id':  participant_id,
        'prolific_pid':    prolific_pid,
        'study_id':        study_id,
        'session_id':      session_id_param,
        'test_mode':       test_mode,
        'start_timestamp': utcnow(),
        'current_step':    'consent',
        'user_agent':      request.headers.get('User-Agent', ''),
        'screen_w':        None,
        'screen_h':        None,
    })
    return redirect(url_for('consent'))


# ── 1. Consent ────────────────────────────────────────────────────────────────
@app.route('/consent', methods=['GET', 'POST'])
def consent():
    if request.method == 'POST':
        p = get_current_participant()
        if not p:
            return redirect(url_for('entry'))
        screen_w = request.form.get('screen_w')
        screen_h = request.form.get('screen_h')
        db_module.update_participant(g.db, p['participant_id'], {
            'consent':           1,
            'consent_timestamp': utcnow(),
            'ts_consent':        utcnow(),
            'current_step':      'demographics',
            'screen_w':          int(screen_w) if screen_w else None,
            'screen_h':          int(screen_h) if screen_h else None,
        })
        return redirect(url_for('demographics'))
    return render_template('consent.html', prog=progress('consent'))


# ── 2. Trait measures ─────────────────────────────────────────────────────────
BPS_ITEMS = [
    "I often find myself at “loose ends,” not knowing what to do.",
    "I find it hard to entertain myself.",
    "Many things I have to do are repetitive and monotonous.",
    "It takes more stimulation to get me going than most people.",
    "I don’t feel motivated by most things that I do.",
    "In most situations, it is hard for me to find something to do or see to keep me interested.",
    "Much of the time, I just sit around doing nothing.",
    "Unless I am doing something exciting, even dangerous, I feel half-dead and dull.",
]
MLQ_ITEMS = [
    # Presence of Meaning subscale (Steger et al., 2006) — items 6–10 of original MLQ
    "I understand my life's meaning.",
    "My life has a clear sense of purpose.",
    "I have a good sense of what makes my life meaningful.",
    "I have discovered a satisfying life purpose.",
    "My life has no clear purpose.",                           # reverse-scored
]
SPA_ITEMS = [
    # Subjective Personal Agency — Yamaguchi et al. (2025) — 1–5
    "I think for myself and make my own life decisions.",
    "I have an idea of what I want to do and/or how I want to be.",
    "I am taking concrete steps to realise what I want to do and/or how I want to be.",
    "I express myself in a way that values my own personal style.",
    "I am able to express my thoughts and feelings in my own words.",
]

BPNS_ITEMS = [
    # BPNS Autonomy Subscale — Deci & Ryan (2000); Johnston & Finney (2010) — 1–7
    "I feel like I am free to decide for myself how to live my life.",
    "I feel pressured in my life.",
    "I generally feel free to express my ideas and opinions.",
    "In my daily life, I frequently have to do what I am told.",
    "People I interact with on a daily basis tend to take my feelings into consideration.",
    "I feel like I can pretty much be myself in my daily situations.",
    "There is not much opportunity for me to decide for myself how to do things in my daily life.",
]

BPNSF_ITEMS = [
    # BPNSF Autonomy Satisfaction & Frustration — Chen et al. (2015) — 1–5
    "I feel a sense of choice and freedom in the things I undertake.",
    "I feel that my decisions reflect what I really want.",
    "I feel my choices express who I really am.",
    "I feel I have been doing what really interests me.",
    'Most of the things I do feel like "I have to".',
    "I feel forced to do many things I wouldn't choose to do.",
    "I feel pressured to do too many things.",
    "My daily activities feel like a chain of obligations.",
]

SE_ITEMS = [
    # General Self-Efficacy — Schwarzer & Jerusalem (1995) — 1–4
    "I can always manage to solve difficult problems if I try hard enough.",
    "If someone opposes me, I can find the means and ways to get what I want.",
    "It is easy for me to stick to my aims and accomplish my goals.",
    "I am confident that I could deal efficiently with unexpected events.",
    "Thanks to my resourcefulness, I know how to handle unforeseen situations.",
    "I can solve most problems if I invest the necessary effort.",
    "I can remain calm when facing difficulties because I can rely on my coping abilities.",
    "When I am confronted with a problem, I can usually find several solutions.",
    "If I am in trouble, I can usually think of a solution.",
    "I can usually handle whatever comes my way.",
]

ATTENTION_CHECK_TRAITS = {
    'bps_attn': {
        'text': 'Please select "4 – Somewhat agree" for this item.',
        'correct': 4,
    }
}


@app.route('/page/2', methods=['GET', 'POST'])
@require_step('ladders')
def ladders():
    p = get_current_participant()
    if request.method == 'POST':
        data = {
            'ts_ladders':            utcnow(),
            'current_step':          'traits',
            # Subjective SES ladders
            'ladder_education':      request.form.get('ladder_education'),
            'ladder_money':          request.form.get('ladder_money'),
            'ladder_job':            request.form.get('ladder_job'),
            # Sociometric status ladders
            'sociometric_respect':   request.form.get('sociometric_respect'),
            'sociometric_admired':   request.form.get('sociometric_admired'),
            'sociometric_important': request.form.get('sociometric_important'),
        }
        condition = assign_condition(g.db, Config)
        data['condition'] = condition
        data['assignment_timestamp'] = utcnow()
        db_module.update_participant(g.db, p['participant_id'], data)
        return redirect(url_for('traits'))

    return render_template('ladders.html', prog=progress('ladders'))


@app.route('/page/3', methods=['GET', 'POST'])
@require_step('traits')
def traits():
    p = get_current_participant()
    if request.method == 'POST':
        data = {'ts_traits': utcnow(), 'current_step': 'writing'}

        data['bps_order']    = request.form.get('bps_order', '')
        data['mlq_order']    = request.form.get('mlq_order', '')
        data['spa_order']    = request.form.get('spa_order', '')
        data['bpns_order']   = request.form.get('bpns_order', '')
        data['bpnsf_order']  = request.form.get('bpnsf_order', '')
        data['se_order']     = request.form.get('se_order', '')
        data['scale_block_order'] = request.form.get('scale_block_order', '')

        for i in range(1, 9):  data[f'bps_{i}']   = request.form.get(f'bps_{i}')
        for i in range(1, 6):  data[f'mlq_{i}']   = request.form.get(f'mlq_{i}')
        for i in range(1, 6):  data[f'spa_{i}']   = request.form.get(f'spa_{i}')
        for i in range(1, 8):  data[f'bpns_{i}']  = request.form.get(f'bpns_{i}')
        for i in range(1, 9):  data[f'bpnsf_{i}'] = request.form.get(f'bpnsf_{i}')
        for i in range(1, 11): data[f'se_{i}']    = request.form.get(f'se_{i}')

        # Attention check
        attn_val = request.form.get('bps_attn')
        data['attention_check_pass'] = 1 if attn_val and int(attn_val) == 4 else 0

        db_module.update_participant(g.db, p['participant_id'], data)
        return redirect(url_for('writing'))

    bps_order   = list(range(1, 9));  random.shuffle(bps_order)
    mlq_order   = list(range(1, 6));  random.shuffle(mlq_order)
    spa_order   = list(range(1, 6));  random.shuffle(spa_order)
    bpns_order  = list(range(1, 8));  random.shuffle(bpns_order)
    bpnsf_order = list(range(1, 9));  random.shuffle(bpnsf_order)
    se_order    = list(range(1, 11)); random.shuffle(se_order)
    blocks = ['bps', 'mlq', 'spa', 'bpns', 'bpnsf', 'se']
    random.shuffle(blocks)

    return render_template('traits.html',
        prog=progress('traits'),
        bps_items=BPS_ITEMS,
        mlq_items=MLQ_ITEMS,
        spa_items=SPA_ITEMS,
        bpns_items=BPNS_ITEMS,
        bpnsf_items=BPNSF_ITEMS,
        se_items=SE_ITEMS,
        bps_order=bps_order,
        mlq_order=mlq_order,
        spa_order=spa_order,
        bpns_order=bpns_order,
        bpnsf_order=bpnsf_order,
        se_order=se_order,
        block_order=blocks,
        attn_check=ATTENTION_CHECK_TRAITS['bps_attn'],
    )


# ── 3. Demographics / SES ─────────────────────────────────────────────────────
@app.route('/page/1', methods=['GET', 'POST'])
@require_step('demographics')
def demographics():
    p = get_current_participant()
    if request.method == 'POST':
        # Multi-select fields arrive as lists
        employment = request.form.getlist('ses_employment')
        race       = request.form.getlist('demo_race')
        data = {
            'ts_demographics':      utcnow(),
            'current_step':         'ladders',
            # Demographics
            'demo_age':             request.form.get('demo_age'),
            'demo_gender':          request.form.get('demo_gender'),
            'demo_relationship':    request.form.get('demo_relationship'),
            'demo_ethnicity':       request.form.get('demo_ethnicity'),
            'demo_race':            ', '.join(race),
            'demo_race_other':      request.form.get('demo_race_other', '').strip(),
            # Objective SES
            'ses_education':        request.form.get('ses_education'),
            'ses_household_income': request.form.get('ses_household_income'),
            'ses_personal_income':  request.form.get('ses_personal_income'),
            'ses_employment':       ', '.join(employment),
            'ses_job_title':        request.form.get('ses_job_title', '').strip(),
        }
        db_module.update_participant(g.db, p['participant_id'], data)
        return redirect(url_for('ladders'))

    return render_template('demographics.html', prog=progress('demographics'))


# ── 4. Writing task ───────────────────────────────────────────────────────────
WRITING_PROMPTS = {
    'meaning': (
        "Please take a few minutes to think about your most important life goals — "
        "the things that feel truly self-defining and intrinsically meaningful to you. "
        "In the box below, write about these goals: what they are, why they matter to you "
        "personally, and what makes your life feel meaningful. "
        "[PLACEHOLDER — replace with final IRB-approved wording.]"
    ),
    'autonomy': (
        "Please think of a recent action or decision that you freely chose — something "
        "you did entirely on your own terms, because you genuinely wanted to, without "
        "external pressure. In the box below, describe what you did, why you chose to do it, "
        "and how it felt to act in that self-endorsed way. "
        "[PLACEHOLDER — replace with final IRB-approved wording.]"
    ),
    'control': (
        "Please describe your typical morning routine in as much detail as possible. "
        "What do you usually do from the moment you wake up until you leave for the day "
        "(or start work)? Include the order of activities, roughly how long each takes, "
        "and any variations between weekdays and weekends. "
        "[PLACEHOLDER — replace with final IRB-approved wording.]"
    ),
}


@app.route('/page/4', methods=['GET', 'POST'])
@require_step('writing')
def writing():
    p = get_current_participant()
    if request.method == 'POST':
        text = request.form.get('writing_text', '').strip()
        time_sec = request.form.get('writing_time_sec')
        db_module.update_participant(g.db, p['participant_id'], {
            'writing_text':    text,
            'writing_time_sec': float(time_sec) if time_sec else None,
            'writing_charcount': len(text),
            'ts_writing':      utcnow(),
            'current_step':    'mancheck',
        })
        return redirect(url_for('mancheck'))

    condition = p['condition']
    prompt = WRITING_PROMPTS.get(condition, WRITING_PROMPTS['control'])
    return render_template('writing.html',
        prog=progress('writing'),
        prompt=prompt,
        min_time=Config.MIN_WRITING_TIME_SEC,
        min_chars=Config.MIN_WRITING_CHARS,
    )


# ── 5. Manipulation check ─────────────────────────────────────────────────────
STATEMEAN_ITEMS = [
    # MLQ-Presence state version (Hunter et al. / Steger et al.) — 1–7
    # Item 5 is reverse-scored (noted for analysis; presented as-is to participants)
    "Right now, I understand my life's meaning.",
    "Right now, my life has a clear sense of purpose.",
    "Right now, I have a good sense of what makes my life meaningful.",
    "Right now, I have discovered a satisfying life purpose.",
    "Right now, my life has no clear purpose.",
]
STATEAUTO_ITEMS = [
    # BPNSFS Autonomy Subscale, state-adapted (Chen et al., 2015) — 1–5
    # Items 1–4 = Autonomy Satisfaction; Items 5–8 = Autonomy Frustration
    # Score the two facets separately; do not combine into one score.
    "Right now, I feel a sense of choice and freedom in what I am doing.",
    "Right now, I feel that my decisions reflect what I really want.",
    "Right now, I feel my choices express who I really am.",
    "Right now, I feel I am doing what really interests me.",
    'Right now, what I am doing feels like something I "have to" do.',
    "Right now, I feel forced to do things I wouldn't choose to do.",
    "Right now, I feel pressured to do too many things.",
    "Right now, my activity feels like a chain of obligations.",
]


@app.route('/page/5', methods=['GET', 'POST'])
@require_step('mancheck')
def mancheck():
    p = get_current_participant()
    if request.method == 'POST':
        data = {'ts_mancheck': utcnow(), 'current_step': 'boring_task'}
        for i in range(1, 6):
            data[f'statemean_{i}'] = request.form.get(f'statemean_{i}')
        for i in range(1, 9):
            data[f'stateauto_{i}'] = request.form.get(f'stateauto_{i}')
        data['mancheck_order'] = request.form.get('mancheck_order', '')
        db_module.update_participant(g.db, p['participant_id'], data)
        return redirect(url_for('boring_task'))

    mean_order = list(range(1, 6))
    auto_order = list(range(1, 9))
    random.shuffle(mean_order)
    random.shuffle(auto_order)
    section_order = ['mean', 'auto']
    random.shuffle(section_order)
    order_str = json.dumps({'mean': mean_order, 'auto': auto_order, 'sections': section_order})
    mean_items = [('statemean', i, STATEMEAN_ITEMS[i - 1]) for i in mean_order]
    auto_items = [('stateauto', i, STATEAUTO_ITEMS[i - 1]) for i in auto_order]
    return render_template('mancheck.html',
        prog=progress('mancheck'),
        mean_items=mean_items,
        auto_items=auto_items,
        section_order=section_order,
        mancheck_order=order_str,
    )


# ── 6. Boring task — transcription ───────────────────────────────────────────
BORING_TASK_TARGET_TEXT = (
    "C451 Standard Test Method for Early Stiffening of Hydraulic Cement (Paste Method)\n"
    "C452/C452M Standard Test Method for Potential Expansion of Portland-Cement Mortars Exposed to Sulfate\n"
    "C457/C457M Standard Test Method for Microscopical Determination of Parameters of the Air-Void System in Hardened Concrete\n"
    "C465 Standard Specification for Processing Additions for Use in the Manufacture of Hydraulic Cements\n"
    "C469/C469M Standard Test Method for Static Modulus of Elasticity and Poisson's Ratio of Concrete in Compression\n"
    "C470/C470M Standard Specification for Molds for Forming Concrete Test Cylinders Vertically\n"
    "C490/C490M Standard Practice for Use of Apparatus for the Determination of Length Change of Hardened Cement Paste, Mortar, and Concrete\n"
    "C494/C494M Standard Specification for Chemical Admixtures for Concrete\n"
    "C495 Standard Test Method for Compressive Strength of Lightweight Insulating Concrete\n"
    "C496/C496M Standard Test Method for Splitting Tensile Strength of Cylindrical Concrete Specimens"
)


def _transcription_accuracy(target: str, submitted: str) -> float:
    """SequenceMatcher similarity after normalising whitespace and case."""
    import re
    from difflib import SequenceMatcher
    def normalise(s):
        return re.sub(r'\s+', ' ', s.strip().lower())
    return round(SequenceMatcher(None, normalise(target), normalise(submitted)).ratio(), 4)


@app.route('/page/6', methods=['GET', 'POST'])
@require_step('boring_task')
def boring_task():
    p = get_current_participant()
    if request.method == 'POST':
        submitted = request.form.get('transcription_text', '').strip()
        accuracy = _transcription_accuracy(BORING_TASK_TARGET_TEXT, submitted)
        data = {
            'ts_boring_task':          utcnow(),
            'current_step':            'outcome',
            'boringtask_duration_sec': request.form.get('time_sec'),
            'transcription_text':      submitted,
            'transcription_accuracy':  accuracy,
            'transcription_charcount': len(submitted),
        }
        db_module.update_participant(g.db, p['participant_id'], data)
        return redirect(url_for('outcome'))

    return render_template('boring_task.html',
        prog=progress('boring_task'),
        min_time=Config.BORING_TASK_DURATION_SEC,
    )


# ── 7. Outcome (MSBS) ─────────────────────────────────────────────────────────
MSBS_ITEMS = [
    # Multidimensional State Boredom Scale — Short Form (Hunter et al., 2015)
    "I am wasting time that would be better spent on something else.",
    "I want something to happen but I'm not sure what.",
    "My mind is wandering.",
    "I am easily distracted.",
    "I feel like I'm sitting around waiting for something to happen.",
    "Time is passing by slower than usual.",
    "I seem to be forced to do things that have no value to me.",
    "I feel bored.",
]


@app.route('/page/7', methods=['GET', 'POST'])
@require_step('outcome')
def outcome():
    p = get_current_participant()
    if request.method == 'POST':
        data = {'ts_outcome': utcnow(), 'current_step': 'debrief'}
        for i in range(1, 9):
            data[f'msbs_{i}'] = request.form.get(f'msbs_{i}')
        db_module.update_participant(g.db, p['participant_id'], data)
        return redirect(url_for('debrief'))

    msbs_order = list(range(1, 9))
    random.shuffle(msbs_order)
    return render_template('outcome.html',
        prog=progress('outcome'),
        items=MSBS_ITEMS,
        order=msbs_order,
    )


# ── 8. Debrief ────────────────────────────────────────────────────────────────
@app.route('/page/8', methods=['GET', 'POST'])
@require_step('debrief')
def debrief():
    p = get_current_participant()
    if request.method == 'POST':
        finish = utcnow()
        start = p['start_timestamp']
        try:
            from datetime import datetime
            start_dt = datetime.fromisoformat(start)
            finish_dt = datetime.fromisoformat(finish)
            duration = (finish_dt - start_dt).total_seconds()
        except Exception:
            duration = None
        db_module.update_participant(g.db, p['participant_id'], {
            'ts_debrief':       finish,
            'finish_timestamp': finish,
            'total_duration_sec': duration,
            'completed':        1,
            'current_step':     'completed',
        })
        session.pop('participant_id', None)
        return redirect(Config.PROLIFIC_COMPLETION_URL)

    return render_template('debrief.html', prog=progress('debrief'))


@app.route('/already-done')
def already_done():
    return render_template('already_done.html')


# ── Admin ─────────────────────────────────────────────────────────────────────
def require_admin(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return wrapped


@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    error = None
    if request.method == 'POST':
        if request.form.get('password') == Config.ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            return redirect(url_for('admin'))
        error = 'Incorrect password.'
    return render_template('admin_login.html', error=error)


@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_login'))


@app.route('/admin')
@require_admin
def admin():
    counts = db_module.get_condition_counts(g.db)
    stats = db_module.get_summary_stats(g.db)
    return render_template('admin.html', counts=counts, stats=stats)


@app.route('/admin/export/xlsx')
@require_admin
def export_xlsx():
    rows = db_module.get_all_participants(g.db)
    counts = db_module.get_condition_counts(g.db)

    wb = openpyxl.Workbook()

    # ── Master sheet ─────────────────────────────────────
    ws = wb.active
    ws.title = 'Data'
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row[h] for h in headers])
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # ── Counts sheet ─────────────────────────────────────
    ws2 = wb.create_sheet('Counts')
    ws2.append(['Condition', 'N assigned', 'N completed'])
    for cond in ['meaning', 'autonomy', 'control']:
        c = counts.get(cond, {'n': 0, 'completed': 0})
        ws2.append([cond, c['n'], c['completed']])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        download_name='boredom_study_data.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/admin/export/db')
@require_admin
def export_db():
    return send_file(
        Config.DATABASE_PATH,
        download_name='study_backup.db',
        as_attachment=True,
        mimetype='application/octet-stream',
    )


@app.route('/admin/delete-all', methods=['POST'])
@require_admin
def delete_all():
    if request.form.get('confirm') != 'DELETE':
        return redirect(url_for('admin') + '?error=Type+DELETE+to+confirm')
    db_module.delete_all_data(g.db)
    return redirect(url_for('admin') + '?deleted=1')


# ── Initialise DB on startup (runs under both gunicorn and python app.py) ─────
db_module.init_db()

if __name__ == '__main__':
    app.run(debug=True, port=5000)
